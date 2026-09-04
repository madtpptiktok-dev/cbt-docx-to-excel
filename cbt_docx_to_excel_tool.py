import re
import os
import shutil
import subprocess
import threading
import queue
import tkinter as tk
from tkinter import messagebox, scrolledtext

import openpyxl
from docx import Document


APP_TITLE = "CBT DOCX → Excel Converter v6"


# =========================================================
# REGEX
# =========================================================

# Contoh:
# 1. She ...
# 6._____ your brother ...
# 15) What ...
QUESTION_START_RE = re.compile(
    r'^\s*(\d{1,3})\s*[\.\)]\s*(.*)$'
)

# HANYA A. / A)
# Sengaja TIDAK menerima "A:" supaya dialog A: / B: tidak dianggap opsi.
OPTION_RE = re.compile(
    r'^\s*([A-Ea-e])\s*[\.\)]\s*(.*)$'
)

# Support:
# Jawaban: B
# Jawaban: B. goes
# Jawabannya : C
# Kunci: D
# Kunci Jawaban : D
# Answer: C
# Answer Key: C
ANSWER_RE = re.compile(
    r'^\s*'
    r'(?:'
        r'Jawaban(?:nya)?'
        r'|Kunci(?:\s+Jawaban(?:nya)?)?'
        r'|Answer(?:\s+Key)?'
    r')'
    r'\s*[:=\-]\s*'
    r'(?P<key>[A-Ea-e]?)'
    r'(?:\s*[\.\)]\s*(?P<label>.*))?'
    r'\s*$',
    re.IGNORECASE
)

SELESAI_RE = re.compile(
    r'^\s*SELESAI\s*$',
    re.IGNORECASE
)

SEPARATOR_RE = re.compile(
    r'^[\s\-_—–=]+$'
)


# =========================================================
# HELPERS
# =========================================================

def clean_text(text):
    return "\n".join(
        line.strip()
        for line in str(text).splitlines()
        if line.strip()
    ).strip()


def normalize_compare(text):
    text = str(text or "").strip()

    text = re.sub(
        r'^[A-Ea-e]\s*[\.\)]\s*',
        '',
        text
    )

    text = re.sub(
        r'\s+',
        ' ',
        text
    )

    return text.strip().lower()


def get_default_dir():
    candidates = [
        os.path.expanduser("~/Downloads"),
        os.path.expanduser("~/Desktop"),
        os.path.expanduser("~/Documents"),
        os.path.expanduser("~"),
    ]

    for path in candidates:
        if os.path.isdir(path):
            return path

    return os.path.expanduser("~")


def native_file_picker(title, extensions):
    """
    Gunakan Zenity di Linux jika tersedia.
    Fallback ke Tkinter.
    """

    initial_dir = get_default_dir()

    if shutil.which("zenity"):

        patterns = " ".join(
            f"*.{ext}"
            for ext in extensions
        )

        label = "/".join(
            ext.upper()
            for ext in extensions
        )

        cmd = [
            "zenity",
            "--file-selection",
            f"--title={title}",
            f"--filename={initial_dir}/",
            f"--file-filter={label} files | {patterns}",
            "--file-filter=All files | *",
        ]

        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True
            )

            if result.returncode == 0:
                return result.stdout.strip()

            return ""

        except Exception:
            pass

    from tkinter import filedialog

    filetypes = []

    if "docx" in extensions:
        filetypes.append(
            ("Word Document", "*.docx")
        )

    if "xlsx" in extensions:
        filetypes.append(
            ("Excel Workbook", "*.xlsx")
        )

    filetypes.append(
        ("All Files", "*.*")
    )

    return filedialog.askopenfilename(
        title=title,
        initialdir=initial_dir,
        filetypes=filetypes
    )


# =========================================================
# DOCX READER
# =========================================================

def docx_to_lines(docx_path):
    """
    Baca paragraph DOCX BERURUTAN.

    Sengaja tidak append semua tabel di akhir,
    karena tabel "Kunci Jawaban" bisa membuat urutan dokumen rusak.
    """

    doc = Document(docx_path)

    lines = []

    for para in doc.paragraphs:

        text = para.text.replace(
            "\r",
            "\n"
        )

        for line in text.splitlines():

            line = line.strip()

            if line:
                lines.append(line)

    return lines


def remove_initial_header(lines):
    """
    Buang header umum:
    SOAL BAHASA INGGRIS
    KELAS : X
    """

    lines = list(lines)

    while lines:

        first = lines[0]

        if re.match(
            r'^\s*SOAL\b',
            first,
            re.IGNORECASE
        ):
            lines.pop(0)
            continue

        if re.match(
            r'^\s*KELAS\s*:',
            first,
            re.IGNORECASE
        ):
            lines.pop(0)
            continue

        break

    return lines


# =========================================================
# PARSER V6
# =========================================================

def parse_docx(docx_path):

    all_lines = docx_to_lines(
        docx_path
    )

    # -----------------------------------------
    # STOP TOTAL SAAT MENEMUKAN "SELESAI"
    # -----------------------------------------

    work_lines = []
    ignored_after_selesai = []
    found_selesai = False

    for index, line in enumerate(all_lines):

        if SELESAI_RE.match(line):

            found_selesai = True

            ignored_after_selesai = (
                all_lines[index + 1:]
            )

            break

        work_lines.append(line)

    work_lines = remove_initial_header(
        work_lines
    )

    # -----------------------------------------
    # STATE MACHINE
    # -----------------------------------------

    data = []

    pending_context = []

    current = None
    current_option = None

    source_question_count = 0
    answer_count = 0

    errors = []
    warnings = []

    for line in work_lines:

        # Abaikan separator panjang ------
        if SEPARATOR_RE.match(line):
            continue

        # ================================
        # 1. NOMOR SOAL
        # ================================

        question_match = QUESTION_START_RE.match(
            line
        )

        if question_match:

            original_number = int(
                question_match.group(1)
            )

            question_text = (
                question_match
                .group(2)
                .strip()
            )

            source_question_count += 1

            # Ada soal baru sebelum soal sebelumnya punya kunci.
            if current is not None:

                errors.append(
                    f"Soal sumber {current['OriginalNomor']} "
                    f"belum mempunyai kunci jawaban sebelum "
                    f"soal {original_number} dimulai."
                )

                # Jangan hilangkan datanya.
                current["Nomor"] = (
                    len(data) + 1
                )

                data.append(current)

            current = {
                "Nomor": 0,
                "OriginalNomor": original_number,
                "Context": clean_text(
                    "\n".join(
                        pending_context
                    )
                ),
                "StemLines": (
                    [question_text]
                    if question_text
                    else []
                ),
                "A": [],
                "B": [],
                "C": [],
                "D": [],
                "E": [],
                "Kunci": "",
                "AnswerLabel": "",
            }

            pending_context = []
            current_option = None

            continue

        # ================================
        # 2. KUNCI JAWABAN
        # ================================

        answer_match = ANSWER_RE.match(
            line
        )

        if (
            answer_match
            and current is not None
        ):

            answer_count += 1

            key = (
                answer_match
                .group("key")
                or ""
            ).upper()

            answer_label = (
                answer_match
                .group("label")
                or ""
            ).strip()

            current["Kunci"] = key
            current["AnswerLabel"] = (
                answer_label
            )

            current["Nomor"] = (
                len(data) + 1
            )

            data.append(current)

            current = None
            current_option = None

            continue

        # ================================
        # 3. PILIHAN A-E
        # ================================

        option_match = OPTION_RE.match(
            line
        )

        if (
            option_match
            and current is not None
        ):

            letter = (
                option_match
                .group(1)
                .upper()
            )

            option_text = (
                option_match
                .group(2)
                .strip()
            )

            current_option = letter

            current[letter].append(
                option_text
            )

            continue

        # ================================
        # 4. TEKS BIASA
        # ================================

        if current is None:

            # Contoh:
            # My Daily Routine
            # paragraf bacaan...
            #
            # Akan ditempel ke soal berikutnya.
            pending_context.append(line)

        else:

            if current_option:

                # Line lanjutan dari opsi.
                current[
                    current_option
                ].append(line)

            else:

                # Line lanjutan stem/dialog.
                current[
                    "StemLines"
                ].append(line)

    # -----------------------------------------
    # DOKUMEN BERAKHIR SAAT SOAL MASIH TERBUKA
    # -----------------------------------------

    if current is not None:

        errors.append(
            f"Soal sumber {current['OriginalNomor']} "
            f"belum ditutup dengan kunci jawaban."
        )

        current["Nomor"] = (
            len(data) + 1
        )

        data.append(current)

    # -----------------------------------------
    # NORMALISASI
    # -----------------------------------------

    result = []

    for index, item in enumerate(
        data,
        start=1
    ):

        context = item[
            "Context"
        ].strip()

        stem = clean_text(
            "\n".join(
                item[
                    "StemLines"
                ]
            )
        )

        if context:

            soal = (
                context
                + "\n"
                + stem
            ).strip()

        else:
            soal = stem

        normalized = {
            "Nomor": index,

            "OriginalNomor":
                item["OriginalNomor"],

            "Soal": soal,

            "A": clean_text(
                "\n".join(
                    item["A"]
                )
            ),

            "B": clean_text(
                "\n".join(
                    item["B"]
                )
            ),

            "C": clean_text(
                "\n".join(
                    item["C"]
                )
            ),

            "D": clean_text(
                "\n".join(
                    item["D"]
                )
            ),

            "E": clean_text(
                "\n".join(
                    item["E"]
                )
            ),

            "Kunci":
                item["Kunci"],

            "AnswerLabel":
                item["AnswerLabel"],
        }

        result.append(
            normalized
        )

    # -----------------------------------------
    # VALIDASI SOURCE
    # -----------------------------------------

    for item in result:

        number = item["Nomor"]

        if not item["Soal"]:

            errors.append(
                f"Soal {number}: "
                f"teks soal kosong."
            )

        for letter in (
            "A",
            "B",
            "C",
            "D",
            "E"
        ):

            if not item[letter]:

                errors.append(
                    f"Soal {number}: "
                    f"pilihan {letter} kosong."
                )

        if item["Kunci"] not in (
            "A",
            "B",
            "C",
            "D",
            "E"
        ):

            warnings.append(
                f"Soal {number}: "
                f"kunci jawaban kosong "
                f"atau tidak valid."
            )

        # Jika source menulis:
        # Jawaban: B. goes
        #
        # pastikan "goes" sama dengan pilihan B.
        elif item["AnswerLabel"]:

            selected_option = item[
                item["Kunci"]
            ]

            if (
                normalize_compare(
                    item["AnswerLabel"]
                )
                !=
                normalize_compare(
                    selected_option
                )
            ):

                warnings.append(
                    f"Soal {number}: "
                    f"teks setelah kunci "
                    f"'{item['AnswerLabel']}' "
                    f"tidak sama dengan "
                    f"pilihan {item['Kunci']} "
                    f"'{selected_option}'."
                )

    # Jumlah nomor soal vs jumlah parser.
    if source_question_count != len(result):

        errors.append(
            f"Jumlah nomor soal sumber "
            f"({source_question_count}) "
            f"tidak sama dengan hasil parser "
            f"({len(result)})."
        )

    # Jumlah kunci vs jumlah soal.
    if answer_count != len(result):

        errors.append(
            f"Jumlah kunci yang dikenali "
            f"({answer_count}) "
            f"tidak sama dengan jumlah soal "
            f"({len(result)})."
        )

    # -----------------------------------------
    # VALIDASI NOMOR ASLI
    # Tidak fatal karena beberapa source memang
    # bisa punya nomor dobel.
    # -----------------------------------------

    original_numbers = [
        item["OriginalNomor"]
        for item in result
    ]

    expected_numbers = list(
        range(
            1,
            len(result) + 1
        )
    )

    if (
        original_numbers
        != expected_numbers
    ):

        warnings.append(
            "Nomor soal di DOCX tidak berurutan "
            "sempurna. Excel tetap akan "
            "dinomori ulang 1 sampai "
            f"{len(result)}."
        )

    # Ada teks sebelum SELESAI yang tidak pernah
    # masuk ke soal berikutnya.
    leftover_context = clean_text(
        "\n".join(
            pending_context
        )
    )

    if leftover_context:

        warnings.append(
            "Ada teks tambahan setelah kunci "
            "soal terakhir sebelum SELESAI:\n"
            + leftover_context[:300]
        )

    meta = {
        "source_question_count":
            source_question_count,

        "answer_count":
            answer_count,

        "parsed_count":
            len(result),

        "found_selesai":
            found_selesai,

        "ignored_after_selesai":
            len(
                ignored_after_selesai
            ),

        "errors":
            errors,

        "warnings":
            warnings,
    }

    return result, meta


# =========================================================
# EXCEL
# =========================================================

def find_target_sheet(wb):

    preferred = (
        "Template Pengisian Soal"
    )

    if preferred in wb.sheetnames:
        return wb[preferred]

    for name in wb.sheetnames:

        low = name.lower()

        if (
            "template" in low
            and
            "soal" in low
        ):
            return wb[name]

    raise ValueError(
        f"Sheet '{preferred}' "
        f"tidak ditemukan.\n\n"
        f"Sheet tersedia:\n"
        + "\n".join(
            wb.sheetnames
        )
    )


def build_excel(
    data,
    template_path,
    output_path
):

    wb = openpyxl.load_workbook(
        template_path
    )

    ws = find_target_sheet(wb)

    start_row = 3

    columns = {
        "Nomor": 1,
        "Jenis": 2,
        "Bobot": 3,
        "Soal": 7,
        "A": 9,
        "B": 10,
        "C": 11,
        "D": 12,
        "E": 13,
        "Kunci": 15,
    }

    # Bersihkan value lama.
    # Formatting template tetap ada.
    for row in range(
        start_row,
        ws.max_row + 1
    ):

        for column in (
            columns.values()
        ):

            ws.cell(
                row=row,
                column=column
            ).value = None

    # Tulis hasil baru.
    for index, item in enumerate(
        data
    ):

        row = (
            start_row
            + index
        )

        ws.cell(
            row=row,
            column=columns["Nomor"],
            value=item["Nomor"]
        )

        ws.cell(
            row=row,
            column=columns["Jenis"],
            value="PG"
        )

        ws.cell(
            row=row,
            column=columns["Bobot"],
            value=10
        )

        ws.cell(
            row=row,
            column=columns["Soal"],
            value=item["Soal"]
        )

        for letter in (
            "A",
            "B",
            "C",
            "D",
            "E"
        ):

            ws.cell(
                row=row,
                column=columns[letter],
                value=item[letter]
            )

        ws.cell(
            row=row,
            column=columns["Kunci"],
            value=item["Kunci"]
        )

    wb.save(
        output_path
    )


def verify_excel(
    data,
    output_path
):

    wb = openpyxl.load_workbook(
        output_path,
        data_only=False
    )

    ws = find_target_sheet(wb)

    start_row = 3

    mismatch = []

    expected_columns = {
        1: "Nomor",
        2: "Jenis",
        3: "Bobot",
        7: "Soal",
        9: "A",
        10: "B",
        11: "C",
        12: "D",
        13: "E",
        15: "Kunci",
    }

    for index, item in enumerate(
        data
    ):

        row = (
            start_row
            + index
        )

        expected = {
            "Nomor":
                item["Nomor"],

            "Jenis":
                "PG",

            "Bobot":
                10,

            "Soal":
                item["Soal"],

            "A":
                item["A"],

            "B":
                item["B"],

            "C":
                item["C"],

            "D":
                item["D"],

            "E":
                item["E"],

            "Kunci":
                item["Kunci"],
        }

        for column, field in (
            expected_columns.items()
        ):

            actual = ws.cell(
                row=row,
                column=column
            ).value

            expected_value = (
                expected[field]
            )

            actual_normalized = (
                ""
                if actual is None
                else str(actual).strip()
            )

            expected_normalized = (
                ""
                if expected_value is None
                else str(
                    expected_value
                ).strip()
            )

            if (
                actual_normalized
                !=
                expected_normalized
            ):

                mismatch.append(
                    f"Soal {item['Nomor']} "
                    f"[{field}] berbeda.\n"
                    f"Source: "
                    f"{expected_normalized!r}\n"
                    f"Excel : "
                    f"{actual_normalized!r}"
                )

    return mismatch


# =========================================================
# GUI
# =========================================================

class App:

    def __init__(
        self,
        root
    ):

        self.root = root

        self.root.title(
            APP_TITLE
        )

        self.root.geometry(
            "880x740"
        )

        self.root.minsize(
            760,
            600
        )

        self.events = queue.Queue()

        self.running = False

        self.docx_var = tk.StringVar()
        self.xlsx_var = tk.StringVar()
        self.output_var = tk.StringVar()

        main = tk.Frame(
            root,
            padx=14,
            pady=14
        )

        main.pack(
            fill="both",
            expand=True
        )

        tk.Label(
            main,
            text=(
                "CBT DOCX → Excel "
                "Converter v6"
            ),
            font=(
                "Arial",
                18,
                "bold"
            )
        ).pack(
            pady=(0, 8)
        )

        self.status = tk.Label(
            main,
            text="Status: Siap",
            anchor="w",
            font=(
                "Arial",
                10,
                "bold"
            )
        )

        self.status.pack(
            fill="x",
            pady=(0, 12)
        )

        self.make_file_row(
            main,
            "File DOCX:",
            self.docx_var,
            self.choose_docx
        )

        self.make_file_row(
            main,
            "Template XLSX:",
            self.xlsx_var,
            self.choose_xlsx
        )

        output_frame = tk.Frame(
            main
        )

        output_frame.pack(
            fill="x",
            pady=7
        )

        tk.Label(
            output_frame,
            text="Nama output:",
            width=16,
            anchor="w"
        ).pack(
            side="left"
        )

        tk.Entry(
            output_frame,
            textvariable=(
                self.output_var
            )
        ).pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        tk.Label(
            output_frame,
            text=".xlsx"
        ).pack(
            side="left"
        )

        self.process_button = tk.Button(
            main,
            text="PROSES & VALIDASI",
            font=(
                "Arial",
                11,
                "bold"
            ),
            command=(
                self.start_process
            ),
            height=2
        )

        self.process_button.pack(
            fill="x",
            pady=14
        )

        tk.Label(
            main,
            text="Log validasi:",
            anchor="w",
            font=(
                "Arial",
                10,
                "bold"
            )
        ).pack(
            fill="x"
        )

        self.log = (
            scrolledtext
            .ScrolledText(
                main,
                wrap="word",
                font=(
                    "Courier New",
                    9
                )
            )
        )

        self.log.pack(
            fill="both",
            expand=True,
            pady=(5, 0)
        )

        self.log.insert(
            "end",
            (
                "v6 siap.\n"
                "Parser menggunakan "
                "nomor soal + opsi A-E "
                "+ kunci.\n"
                "Aplikasi berhenti membaca "
                "saat menemukan SELESAI.\n"
            )
        )

        self.root.after(
            100,
            self.poll_events
        )

        self.root.protocol(
            "WM_DELETE_WINDOW",
            self.on_close
        )

    def make_file_row(
        self,
        parent,
        label,
        variable,
        command
    ):

        frame = tk.Frame(
            parent
        )

        frame.pack(
            fill="x",
            pady=7
        )

        tk.Label(
            frame,
            text=label,
            width=16,
            anchor="w"
        ).pack(
            side="left"
        )

        tk.Entry(
            frame,
            textvariable=variable
        ).pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 8)
        )

        tk.Button(
            frame,
            text="Pilih File",
            command=command
        ).pack(
            side="left"
        )

    def choose_docx(self):

        if self.running:
            return

        path = native_file_picker(
            "Pilih file DOCX",
            ["docx"]
        )

        if path:

            self.docx_var.set(
                path
            )

            base = (
                os.path.splitext(
                    os.path.basename(
                        path
                    )
                )[0]
            )

            self.output_var.set(
                base
                + "_CBT_READY"
            )

    def choose_xlsx(self):

        if self.running:
            return

        path = native_file_picker(
            "Pilih template XLSX",
            ["xlsx"]
        )

        if path:
            self.xlsx_var.set(
                path
            )

    def emit(
        self,
        kind,
        value=None
    ):

        self.events.put(
            (
                kind,
                value
            )
        )

    def start_process(self):

        if self.running:
            return

        docx_path = (
            self.docx_var
            .get()
            .strip()
        )

        xlsx_path = (
            self.xlsx_var
            .get()
            .strip()
        )

        output_name = (
            self.output_var
            .get()
            .strip()
        )

        if not os.path.isfile(
            docx_path
        ):

            messagebox.showerror(
                "Error",
                "Pilih file DOCX "
                "yang valid."
            )

            return

        if not os.path.isfile(
            xlsx_path
        ):

            messagebox.showerror(
                "Error",
                "Pilih template XLSX "
                "yang valid."
            )

            return

        if not output_name:

            output_name = (
                "SOAL_CBT_READY"
            )

        if not output_name.lower().endswith(
            ".xlsx"
        ):

            output_name += (
                ".xlsx"
            )

        output_path = os.path.join(
            os.path.dirname(
                docx_path
            ),
            output_name
        )

        self.running = True

        self.process_button.config(
            state="disabled",
            text="SEDANG MEMPROSES..."
        )

        self.status.config(
            text="Status: Memproses..."
        )

        self.log.delete(
            "1.0",
            "end"
        )

        worker_thread = threading.Thread(
            target=self.worker,
            args=(
                docx_path,
                xlsx_path,
                output_path
            ),
            daemon=True
        )

        worker_thread.start()

    def worker(
        self,
        docx_path,
        xlsx_path,
        output_path
    ):

        try:

            self.emit(
                "log",
                "1/4 Membaca + parsing DOCX...\n"
            )

            data, meta = parse_docx(
                docx_path
            )

            self.emit(
                "log",
                (
                    f"Nomor soal sumber     : "
                    f"{meta['source_question_count']}\n"

                    f"Soal berhasil diparse : "
                    f"{meta['parsed_count']}\n"

                    f"Kunci dikenali        : "
                    f"{meta['answer_count']}\n"

                    f"Marker SELESAI        : "
                    f"{'DITEMUKAN' if meta['found_selesai'] else 'TIDAK ADA'}\n"

                    f"Baris setelah SELESAI : "
                    f"{meta['ignored_after_selesai']} "
                    f"(diabaikan)\n\n"
                )
            )

            self.emit(
                "log",
                "2/4 Validasi sumber...\n"
            )

            errors = (
                meta["errors"]
            )

            warnings = (
                meta["warnings"]
            )

            if errors:

                for error in errors:

                    self.emit(
                        "log",
                        "ERROR: "
                        + error
                        + "\n"
                    )

            if warnings:

                for warning in warnings:

                    self.emit(
                        "log",
                        "WARNING: "
                        + warning
                        + "\n"
                    )

            if (
                not errors
                and
                not warnings
            ):

                self.emit(
                    "log",
                    (
                        "OK: Struktur sumber "
                        "valid.\n"
                    )
                )

            self.emit(
                "log",
                (
                    f"\nError   : "
                    f"{len(errors)}\n"

                    f"Warning : "
                    f"{len(warnings)}\n\n"
                )
            )

            # ==================================
            # JANGAN BUAT EXCEL JIKA SOURCE RUSAK
            # ==================================

            if errors:

                raise ValueError(
                    (
                        f"Ditemukan "
                        f"{len(errors)} "
                        f"ERROR parsing.\n\n"
                        f"Excel TIDAK dibuat "
                        f"supaya tidak menghasilkan "
                        f"file setengah jadi."
                    )
                )

            # ==================================
            # WRITE
            # ==================================

            self.emit(
                "log",
                (
                    "3/4 Menulis ke "
                    "template Excel...\n"
                )
            )

            build_excel(
                data,
                xlsx_path,
                output_path
            )

            # ==================================
            # VERIFY
            # ==================================

            self.emit(
                "log",
                (
                    "4/4 Membandingkan "
                    "hasil Excel dengan parser...\n"
                )
            )

            mismatch = verify_excel(
                data,
                output_path
            )

            self.emit(
                "log",
                "\n=== HASIL FINAL ===\n"
            )

            self.emit(
                "log",
                (
                    f"Soal sumber : "
                    f"{meta['source_question_count']}\n"

                    f"Soal Excel  : "
                    f"{len(data)}\n"

                    f"Kunci       : "
                    f"{meta['answer_count']}\n"

                    f"Error       : "
                    f"{len(errors)}\n"

                    f"Warning     : "
                    f"{len(warnings)}\n"

                    f"Mismatch    : "
                    f"{len(mismatch)}\n"
                )
            )

            if mismatch:

                self.emit(
                    "log",
                    "\nPERBEDAAN:\n"
                )

                for item in mismatch:

                    self.emit(
                        "log",
                        item + "\n\n"
                    )

            else:

                self.emit(
                    "log",
                    (
                        "\nPASS: DOCX → parser → "
                        "Excel cocok.\n"
                    )
                )

            self.emit(
                "log",
                (
                    "\nOutput:\n"
                    + output_path
                    + "\n"
                )
            )

            self.emit(
                "done",
                {
                    "output":
                        output_path,

                    "count":
                        len(data),

                    "warnings":
                        len(warnings),

                    "mismatch":
                        len(mismatch),

                    "found_selesai":
                        meta[
                            "found_selesai"
                        ],
                }
            )

        except Exception as error:

            self.emit(
                "error",
                str(error)
            )

    def poll_events(self):

        try:

            while True:

                kind, value = (
                    self.events
                    .get_nowait()
                )

                if kind == "log":

                    self.log.insert(
                        "end",
                        value
                    )

                    self.log.see(
                        "end"
                    )

                elif kind == "done":

                    self.running = False

                    self.process_button.config(
                        state="normal",
                        text="PROSES & VALIDASI"
                    )

                    if (
                        value["mismatch"]
                        == 0
                    ):

                        self.status.config(
                            text=(
                                "Status: PASS"
                            )
                        )

                    else:

                        self.status.config(
                            text=(
                                "Status: "
                                "Selesai dengan masalah"
                            )
                        )

                    messagebox.showinfo(
                        "Selesai",
                        (
                            f"Total soal: "
                            f"{value['count']}\n"

                            f"Warning: "
                            f"{value['warnings']}\n"

                            f"Mismatch: "
                            f"{value['mismatch']}\n\n"

                            f"Output:\n"
                            f"{value['output']}"
                        )
                    )

                elif kind == "error":

                    self.running = False

                    self.process_button.config(
                        state="normal",
                        text="PROSES & VALIDASI"
                    )

                    self.status.config(
                        text="Status: ERROR"
                    )

                    self.log.insert(
                        "end",
                        (
                            "\nFATAL / STOP:\n"
                            + value
                            + "\n"
                        )
                    )

                    self.log.see(
                        "end"
                    )

                    messagebox.showerror(
                        "Proses Dihentikan",
                        value
                    )

        except queue.Empty:
            pass

        try:

            self.root.after(
                100,
                self.poll_events
            )

        except tk.TclError:
            pass

    def on_close(self):

        self.root.destroy()


if __name__ == "__main__":

    root = tk.Tk()

    App(root)

    root.mainloop()

